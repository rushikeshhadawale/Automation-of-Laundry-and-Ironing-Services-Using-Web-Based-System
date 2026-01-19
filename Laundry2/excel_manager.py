from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

def save_order_to_excel(customer_name, service_type, price):
    file_name = "laundry_data.xlsx"

    # Check if file exists
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Create headers
        sheet.append(["Customer Name", "Service Type", "Price", "Date"])

    # Append the new data
    sheet.append([customer_name, service_type, price, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    # Save the Excel file
    workbook.save(file_name)
    print("✅ Order saved successfully!")
